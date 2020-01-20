import logging
import re
import os
import yaml
import pandas as pd
from fddc.regex import parse_regex
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import get_column_letter

logger = logging.getLogger('spreadsheetcleaner')


# Functions to categorise data

def make_category(config):
    def categorize(series):
        for c in config:
            if series == c['code']:
                return c['code']
            elif c['code'] in str(series):
                return c['code']
            elif c['name'] in str(series):
                return c['code']

            for r in c.get('regex',[]):
                p = parse_regex(r)
                if p.match(str(series)) is not None:
                    return c['code']
                
        return 'not matched'
    return categorize

def make_map(original, new):
    '''
    Creates map of {original value:new value}
    '''
    values = dict(zip(original, new))
    return values


# Function to create data validation object

def validation_from_list(validation_range):
    '''Creates a data validation object based on validation_range (cell range to point to). 
    Requires import of DataValidation from openpyxl.worksheet.datavalidation '''
    from openpyxl.worksheet.datavalidation import DataValidation
    validation = DataValidation(type='list', formula1=validation_range)
    validation.error = 'Your entry is not in the list'
    validation.errorTitle = 'Invalid Entry'
    validation.prompt = 'Please select from the list'
    validation.promptTitle = 'List Selection'
    
    return validation


# Main function to go through spreadsheet and replace data

def clean(input_file, output_file, matching_report_file, data_config, **args):
    '''Replaces values in spreadsheet by standardized values following rules in data_config. 
    Saves clean spreadsheet in clean_path and matching report in matching_path'''
    
    # Set up two workbooks to write both clean spreadsheet and matching report in unique spreadsheets
    writer_clean = pd.ExcelWriter(output_file, engine='xlsxwriter') # create object to store clean data
    wb = Workbook() # create object to store matching report
    ws_ref = wb.active
    ws_ref.title = "References" # This sheet will hold the validation references
    reference_count = 0 #keep track of the columns filled with validation references
    references = pd.DataFrame()
    
    # Run through sheets within spreadsheet (matching items in data_config)
    for item in data_config:
        data = pd.read_excel(input_file, sheet_name=item)
        df = data.copy()
        settings = data_config[item]
        # Create new sheet in matching report
        ws = wb.create_sheet(item) 
        # Create header
        ws['A1'] = 'column'
        ws['B1'] = 'former_value'
        ws['C1'] = 'new_value'
        
        # Run through config criteria for that sheet
        for col in settings:
            if df[df[col].notnull()].shape[0] > 0: # Only look at non-empty columns

                # Map current value to match value
                categorize = make_category(settings[col])
                df[col] = df[col].str.strip()
                unique = df[col].unique()
                clean_series = pd.Series(unique).apply(categorize).tolist()
                match = make_map(unique, clean_series)

                # Replace values in sheet
                df[col] = df[col].replace(match)

                # Write combination of former - new values into match_report
                match_report = pd.DataFrame({'former_value': list(match.keys()), 'new_value': list(match.values())})
                match_report['column'] = col
                match_report = match_report[['column', 'former_value', 'new_value']]
                
                # Write data validation options in 'References' worksheet ws_ref
                reference_count += 1
                options = [c["code"] for c in settings[col]]
                column_name = '{} - {}'.format(item, col)
                if reference_count == 1:
                    references[column_name] = options
                else:
                    new_reference = pd.DataFrame({column_name:options})
                    references = pd.concat([references, new_reference], axis=1)
                
                # Insert match_report dataframe to current worksheet
                for r in dataframe_to_rows(match_report, index=False, header=None):
                    ws.append(r)

                # Create data validation object
                validation_range = "References!${}$2:${}${}".format(get_column_letter(reference_count), 
                                                                    get_column_letter(reference_count), 
                                                                    len(options)+1)
                dv = validation_from_list(validation_range)
                ws.add_data_validation(dv)
                
                # Add cells from match_report to the data validation object
                # Find position of first and last cells that will get the data validation rule in column C
                last_cell = 'C{}'.format(len(ws['C']))
                first_cell = 'C{}'.format(len(ws['C']) - len(match) + 1)
                # Add cells
                dv.add('{}:{}'.format(first_cell, last_cell))
               
        # Save cleaned sheet into excel sheet
        df.to_excel(writer_clean, sheet_name=item, index=False)

    for r in dataframe_to_rows(references, index=False, header=True):
        ws_ref.append(r)
    # Save both reports
    wb.save(matching_report_file)
    writer_clean.save()