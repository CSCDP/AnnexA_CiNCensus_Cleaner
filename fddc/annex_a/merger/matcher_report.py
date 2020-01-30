import logging
import re
from typing import Union, List, Iterable, Tuple, Dict, Any
import dacite
import pandas as pd
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from fddc.annex_a.merger import file_scanner, workbook_util, matcher
from fddc.annex_a.merger.matcher import MatchedColumn, MatchedSheet, SheetWithHeaders
from fddc.annex_a.merger.configuration import SourceConfig
from fddc.annex_a.merger.file_scanner import ScanSource, FileSource
from fddc.annex_a.merger.workbook_util import WorkSheetDetail

logger = logging.getLogger('fddc.annex_a.merger.matcher_report')


@dataclass
class MatchInput:
    filename: str
    sort_key: Any
    header_starts: Any
    sheetname: Union[str, None]
    table: Union[str, None]
    column_name: Union[str, None]
    header_name: Union[str, None]


@dataclass
class __MatchReport:
    records: List[MatchInput]


def __check_nan(value):
    if pd.isnull(value):
        return None
    return value


def parse_report(report: Union[str, pd.DataFrame]) -> List[MatchInput]:
    if isinstance(report, str):
        logger.info(f"Reading match configuration from '{report}'")
        report = pd.read_excel(report)

    assert isinstance(report, pd.DataFrame), f"Expected the report to be a DataFrame, not a {type(report)}"

    # We don't trust users (or even our report) to not create duplicates...
    report.drop_duplicates(subset=['filename', 'sheetname', 'header_name'], keep='first', inplace=True)

    if "header_starts" not in report:
        report["header_starts"] = None

    report = report[[
        "filename",
        "sort_key",
        "header_starts",
        "sheetname",
        "table",
        "column_name",
        "header_name",
    ]]

    report_records = report.to_dict(orient="records")

    match_input = dacite.from_dict(
        data_class=__MatchReport,
        data=dict(records=report_records),
        config=dacite.Config(strict=True, type_hooks={str: __check_nan, int: __check_nan})
    ).records

    return match_input


def process_report(
        match_input: Union[Iterable[MatchInput], pd.DataFrame, str],
        data_sources: List[SourceConfig]
):
    if isinstance(match_input, str) or isinstance(match_input, pd.DataFrame):
        match_input = parse_report(match_input)


    files_to_scan = set()
    columns_per_table = dict()
    mapping_dict: Dict[Tuple, List[MatchInput]] = dict()

    for input in match_input:
        if input.sheetname is None:
            # First we look for unscanned files
            sort_key = None
            if input.sort_key is not None:
                sort_key = [f'/.*/{input.sort_key}/']
            scan_source = ScanSource(include=input.filename, sort_keys=sort_key)
            files = file_scanner.find_input_files(scan_source)
            files_to_scan.update(files)
        else:
            # Then we build a lookup of files and tables to see if any tables have no columns listed
            key = (input.filename, input.sort_key, input.sheetname, input.table)
            columns = columns_per_table.setdefault(key, set())
            if input.column_name is not None:
                columns.add(input.column_name)

        if input.column_name is not None and input.header_name is not None:
            key = (input.filename, input.sort_key, input.sheetname, input.table)
            mapping_dict.setdefault(key, []).append(input)

    matched_list: List[MatchedSheet] = []
    unmatched_list: List[WorkSheetDetail] = []

    for file in files_to_scan:
        worksheets = workbook_util.find_worksheets(file)
        # Match datasources based on configuration
        matched, unmatched = matcher.match_data_sources(worksheets, data_sources)
        matched_list += matched
        unmatched_list += unmatched

    for key, columns in columns_per_table.items():
        if len(columns) == 0:
            file, sort_key, sheetname, table = key
            if table is not None:
                worksheet_list = workbook_util.find_worksheets(FileSource(file, sort_key=sort_key))
                worksheet = next(iter([w for w in worksheet_list if w.sheetname == sheetname]))
                source_config = next(iter([d for d in data_sources if d.name == table]))
                matched = MatchedSheet(sheet_detail=worksheet, source_config=source_config)
                matched_list.append(matched)

    # Match headers to column configuration
    sheet_with_headers: List[SheetWithHeaders] = matcher.match_columns(matched_list)

    for key, mapping_list in mapping_dict.items():
        file, sort_key, sheetname, table = key
        worksheet_list = workbook_util.find_worksheets(FileSource(file, sort_key=sort_key))
        sheet_detail = next(iter([w for w in worksheet_list if w.sheetname == sheetname]))
        source_config = next(iter([d for d in data_sources if d.name == table]))
        sheet = MatchedSheet(sheet_detail=sheet_detail, source_config=source_config)

        column_list: List[MatchedColumn] = []

        for mapping in mapping_list:
            column_config = next(iter([c for c in source_config.columns if c.name == mapping.column_name]))
            header_config = next(iter([h for h in sheet_detail.headers if h.value == mapping.header_name]))
            column = MatchedColumn(column=column_config, header=header_config)
            column_list.append(column)

        sheet_with_headers.append(SheetWithHeaders(sheet=sheet, columns=column_list, unmatched_columns=[]))

    return sheet_with_headers, unmatched_list


def column_report(sheet_list: List[SheetWithHeaders],
                  unmatched_list: List[WorkSheetDetail] = None,
                  filename: str = None) -> pd.DataFrame:
    report: List[Any] = []

    if unmatched_list:
        for sheet_detail in unmatched_list:
            source_info = dict(
                filename=sheet_detail.filename,
                sort_key=sheet_detail.sort_key,
                sheetname=sheet_detail.sheetname,
                header_starts=sheet_detail.header_row_index,
                header_pos='"{0}"'.format('","'.join(sheet_detail.header_names()))
            )
            non_blank_headers = [h for h in sheet_detail.header_names() if len(h) > 0]
            if len(non_blank_headers) == 0:
                report.append(source_info)
            else:
                for header in non_blank_headers:
                    loop_info = dict(source_info)
                    loop_info["header_name"] = header
                    report.append(loop_info)

    for source in sheet_list:
        # Basic information
        source_info = dict(
            filename=source.sheet.sheet_detail.filename,
            sort_key=source.sheet.sheet_detail.sort_key,
            sheetname=source.sheet.sheet_detail.sheetname,
            table=source.sheet.source_config.name,
            header_starts=source.sheet.sheet_detail.header_row_index,
            header_pos='"{0}"'.format('","'.join(source.sheet.sheet_detail.header_names()))
        )

        for col in source.sheet.source_config.columns:
            loop_info = dict(source_info)
            loop_info["column_name"] = col.name
            loop_info["header_name"] = next(iter([h.header.value
                                                  for h in source.columns
                                                  if h.column.name == col.name]), "")
            report.append(loop_info)

        for header in source.unmatched_headers():
            loop_info = dict(source_info)
            loop_info["header_name"] = header.value
            report.append(loop_info)

    df = pd.DataFrame(report)
    df.sort_values(by=["sort_key", "sheetname"], inplace=True)
    df["unmatched"] = None

    if filename is not None:
        df = df[
            ["filename", "sort_key", "header_starts", "sheetname", "table", "column_name",
             "header_name", "header_pos", "unmatched"]]

        sheet_name = "MatchReport"

        logger.info("Writing matching report to ${filename}.")
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        writer.book = Workbook()
        sheet = writer.book.active
        sheet.title = sheet_name

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

        df.to_excel(writer, sheet_name, index=False)

        # Create data table
        data_range = f"A1:{get_column_letter(df.shape[1])}{df.shape[0] + 1}"
        table = Table(displayName=sheet_name, ref=data_range)
        # Adjust column widths

        widths = [10, 10, 5, 20, 20, 30, 30, 15, 15]
        for c in range(df.shape[1]):
            sheet.column_dimensions[get_column_letter(c + 1)].width = widths[c]

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)

        # Insert formula for header position
        for row in sheet.iter_rows(min_row=2, min_col=8, max_col=9):
            if str(row[0].value).startswith('"'):
                row[0].value = re.sub(r"\s+", "", '''
=IF
(
    ISBLANK
    ( 
        [[header_name]]
    ), 
    "",
    IFERROR
    (
        MATCH
        ( 
            [[header_name]] , {{{}}} , 0
        )-1,
        "INVALID"
    )
)
''').format(row[0].value)

            row[1].value = re.sub(r"\s+", "", '''
=IF
(
    ISBLANK
    (
        [[table]]
    ),
    "TABLE",
    IF
    (
        ISBLANK
        (
            [[header_name]]
        ), 
        IF
        ( 
            ISBLANK
            (
                [[column_name]]
            ),
            "",
            "HEADER"
        ),
        IF
        (
            ISBLANK
            (
                [[column_name]]
            ),
            "COLUMN",
            ""
        )
    )
)
''')
        writer.save()

    return df[["filename", "sort_key", "header_starts", "sheetname", "table", "column_name", "header_name"]]