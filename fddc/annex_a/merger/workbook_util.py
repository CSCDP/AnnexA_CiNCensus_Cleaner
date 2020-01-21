import logging

logger = logging.getLogger('fddc.annex_a.merger.workbook_util')


def find_worksheets(filename, **sourceinfo):
    if filename.endswith(".xlsx"):
        return find_worksheets_xlsx(filename=filename, **sourceinfo)
    elif filename.endswith(".xls"):
        return find_worksheets_xls(filename=filename, **sourceinfo)


def find_worksheets_xlsx(filename, **sourceinfo):
    from openpyxl import load_workbook
    from zipfile import BadZipFile

    data_sources = []
    logger.debug("Opening {}".format(filename))
    try:
        workbook = load_workbook(filename=filename, read_only=True)
    except BadZipFile:
        logger.warning("Failed to open Excel file: {}".format(filename), exc_info=False)
        logger.debug("BadZipFile encountered while opening XLSX file", exc_info=True)
        return data_sources

    for sheetname in workbook.sheetnames:
        logger.debug("Checking sheet {} in {}".format(sheetname, filename))
        sheet = workbook[sheetname]

        # We search for first row with more than 3 non-null values
        header_row_index = None
        header_values = []
        for row in sheet.iter_rows(max_row=5):
            row_length = 0
            for col in row:
                if col.value is not None:
                    header_row_index = col.row
                    row_length += 1
            if row_length > 3:
                header_values = [col.value for col in row]
                break

        data_sources.append({
            "filename": filename,
            **sourceinfo,
            "sheetname": sheetname,
            "header_row_index": header_row_index,
            "header_values": header_values
        })

    return data_sources


def find_worksheets_xls(filename, **sourceinfo):
    from xlrd import open_workbook

    data_sources = []
    logger.debug("Opening {}".format(filename))
    workbook = open_workbook(filename=filename)

    for sheet_name in workbook.sheet_names():
        logger.debug("Checking sheet {} in {}".format(sheet_name, filename))
        sheet = workbook.sheet_by_name(sheet_name)

        # We search for first row with more than 3 non-null values
        header_row_index = None
        header_values = []
        for ix, row in enumerate(sheet.get_rows()):
            row_length = 0
            for col in row:
                if col.value is not None and len(col.value.strip()) > 0:
                    header_row_index = ix + 1
                    row_length += 1
            if row_length > 3:
                header_values = [col.value for col in row]
                break

        data_sources.append({
            "filename": filename,
            **sourceinfo,
            "sheetname": sheet_name,
            "header_row_index": header_row_index,
            "header_values": header_values
        })

    return data_sources