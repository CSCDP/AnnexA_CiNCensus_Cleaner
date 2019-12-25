import copy
import glob
import logging
import re
import os
from openpyxl import load_workbook

from fddc.config import ConfigError
from fddc.regex import parse_regex, substitute
from zipfile import BadZipFile

logger = logging.getLogger('fddc.annex_a.merger')


class RegexMatcher():
    def __init__(self, pattern):
        self.__matcher = parse_regex(pattern)

    def match(self, string):
        if self.__matcher.match(string) is not None:
            return True
        else:
            return False


def find_input_files(root, include, sort_keys=None, **args):
    """
    Processes a single item in the input config.
    """

    # Build complete globbing path based on root and include pattern
    root = os.path.abspath(root)
    file_glob = os.path.join(root, include)

    logger.debug("Resolving files using {}".format(file_glob))

    # Search for files and build absolute paths to the files
    files = glob.glob(file_glob, recursive=True)
    files = [os.path.abspath(file) for file in files]

    output = []
    for filename in files:
        sourcename = os.path.relpath(filename, root)

        # Build sort-keys
        if sort_keys is None:
            sort_key = sourcename
        else:
            sort_key = sourcename
            for sk in sort_keys:
                sort_key = substitute(sk, sort_key, sort_key)

        output.append({
            "filename": filename,
            "sourcename": sourcename,
            "sort_key": sort_key,
            "sort_keys": sort_keys,
            "root": root,
            "input_cfg": {
                "root": root,
                "include": include,
                **args
            }})

    return output


def find_worksheets(filename, **sourceinfo):
    data_sources = []

    logger.debug("Opening {}".format(filename))
    try:
        workbook = load_workbook(filename=filename, read_only=True)
    except BadZipFile:
        logger.warning("Failed to open Excel file: {}".format(filename), exc_info=False)
        logger.debug("BadZipFile encountered while opening XLSX file", exc_info=True)
        return data_sources

    for sheetname in workbook.sheetnames:
        logger.debug("Checking sheet {} in {}".format(sheetname, filename))
        sheet = workbook[sheetname]

        # We search for first row with more than 3 non-null values
        header_row_index = None
        header_values = []
        for row in sheet.iter_rows(max_row=5):
            row_length = 0
            for col in row:
                if col.value is not None:
                    header_row_index = col.row
                    row_length += 1
            if row_length > 3:
                header_values = [col.value for col in row]
                break

        data_sources.append({
            "filename": filename,
            **sourceinfo,
            "sheetname": sheetname,
            "header_row_index": header_row_index,
            "header_values": header_values
        })

    return data_sources


def init_datasource_config(datasource_config):
    """
    Reads all configured datasources from the configuration
    and initialises them with patterns.

    Returns a new object.
    """
    datasource_config = copy.deepcopy(datasource_config)
    for key, cfg_source in datasource_config.items():
        # We generate a default regex to use if no specific pattern is set
        name = cfg_source["name"]
        name = re.sub(r'\s+', "\\\s+", name)

        # If regex is not set, use our default one instead
        pattern = cfg_source.get("regex", "/.*{}.*/i".format(name))

        # Set matchers object
        cfg_source["matchers"] = [{"type": "regex", "pattern": pattern}]

    return datasource_config


def init_all_column_config(datasources):
    datasources = copy.deepcopy(datasources)
    for ds in datasources:
        ds_key = ds["source_key"]
        custom_config = ds["input_cfg"].get("datasources", {}).get(ds_key, dict(columns=[]))["columns"]
        global_config = ds["source_config"]["columns"]

        ds["column_config"] = init_column_config(global_config, custom_config)

    return datasources


def init_column_config(global_config, custom_config):
    """
    Iterates over the columns and adds matcher configuration

    Returns merged list
    """
    custom_config_dict = {d["name"]: d for d in custom_config}

    global_config = copy.deepcopy(global_config)

    for col in global_config:
        # We generate a default regex to use if no specific pattern is set
        name = col["name"]
        escaped_name = re.sub(r'\s+', "\\\s+", name)

        # Create default matchers for column
        pattern = col.get("regex", "/.*{}.*/i".format(escaped_name))
        matchers = [{"type": "regex", "pattern": pattern}]

        # See if we have any custom matchers
        custom_col_config = custom_config_dict.get(name)
        if custom_col_config is not None and "regex" in custom_col_config:
            matchers = [{"type": "regex", "pattern": custom_col_config["regex"]}] + matchers

        col["matchers"] = matchers

    return global_config


def get_matcher(type, **args):
    if type == "regex":
        return RegexMatcher(**args)
    else:
        raise Exception("Unknown matcher type: {}".format(type))


def match_datasources(data_sources, datasource_config):
    data_sources = copy.deepcopy(data_sources)
    for source in data_sources:
        for key, cfg_source in datasource_config.items():
            for matcher in cfg_source["matchers"]:
                matcher = get_matcher(**matcher)
                if matcher.match(source["sheetname"]):
                    source["source_key"] = key
                    source["source_config"] = cfg_source
                    break

        if "source_key" not in source:
            logger.warning("No datasource identified for '{}' in '{}'".format(source["sheetname"], source["sourcename"]))
        else:
            logger.debug("Matched datasource {source_key} for sheet {sheetname} in {filename}".format(**source))

    return data_sources


def match_columns(data_sources):
    data_sources = copy.deepcopy(data_sources)
    for ds in data_sources:

        # Build list of available headers
        ds_headers = [{"name": x, "pos": ix} for ix, x in enumerate(ds["header_values"]) if x is not None]

        # Loop over each configured column and try to identify candidate column
        columns = []
        for col in ds["column_config"]:
            col = copy.deepcopy(col)
            columns.append(col)
            for matcher in col["matchers"]:
                matcher = get_matcher(**matcher)
                for header in ds_headers:
                    if matcher.match(header["name"]):
                        ds_headers.remove(header)
                        col["header"] = header
                        break

        ds["column_spec"] = columns
        ds["headers_unmatched"] = ds_headers

    return data_sources


def column_report(data_sources, all_data_sources, filename=None):
    import pandas as pd
    report = []

    invalid_data_sources = [d for d in all_data_sources if "source_key" not in d]
    for source in invalid_data_sources:
        source_info = dict(root=source["root"], filename=source["sourcename"],
                           sort_key=source["sort_key"], sheetname=source["sheetname"])
        report.append(source_info)

    for source in data_sources:
        source_info = dict(root=source["root"], filename=source["sourcename"],
                           sort_key=source["sort_key"], sheetname=source["sheetname"])
        source_info["name"] = source["source_config"]["name"]

        for col in source.get("column_spec", []):
            loop_info = copy.deepcopy(source_info)
            loop_info["column_name"] = col["name"]

            if "header" in col:
                loop_info["header_name"] = col["header"]["name"]
                loop_info["header_pos"] = col["header"]["pos"]

            report.append(loop_info)

        for header in source.get("headers_unmatched", []):
            loop_info = copy.deepcopy(source_info)
            loop_info["header_name"] = header["name"]
            loop_info["header_pos"] = header["pos"]
            report.append(loop_info)

    df = pd.DataFrame(report)
    df.sort_values(by=["sort_key", "sheetname"], inplace=True)
    df.loc[df["column_name"].isnull() & df["header_name"].notnull(), "potential_mismatch"] = True

    if filename is not None:
        df = df[
            ["root", "filename", "sort_key", "sheetname", "name", "column_name", "header_name", "potential_mismatch"]]
        df.to_excel(filename, index=False)

    return df


def load_dataframes(valid_datasources, datasources, **args):
    import pandas as pd
    ds_lookup = {ds["filename"]: {} for ds in valid_datasources}
    for ds in valid_datasources:
        ds_lookup[ds["filename"]][ds["sheetname"]] = copy.deepcopy(ds)

    valid_datasources = []
    for filename in ds_lookup:
        logger.info("Opening {}".format(filename))
        file = pd.ExcelFile(filename)
        for sheetname in ds_lookup[filename]:
            ds = ds_lookup[filename][sheetname]
            valid_datasources.append(ds)
            skiprows = ds.get("header_row_index")
            if skiprows is None:
                skiprows = 1
            logger.debug("Reading '{}' from '{}'".format(sheetname, filename))
            df = pd.read_excel(filename, skiprows=skiprows - 1, sheet_name=sheetname)
            ds["dataframe"] = df
            logger.debug("Reading {} rows and {} cols from '{}' from '{}'".format(*df.shape, sheetname, filename))

    return valid_datasources


def merge_datasources(valid_datasources, datasources, error_report_filename=False, **args):
    import pandas as pd
    import numpy as np

    cfg_datasources = datasources

    output_dataframes = {}

    error_report = []

    for ds_key in cfg_datasources.keys():
        # Find datasources for this key
        datasources = [ds for ds in valid_datasources if ds["source_key"] == ds_key]

        # Find configuration for this datasource
        cfg_datasource = cfg_datasources[ds_key]
        cfg_columns = cfg_datasource["columns"]
        columns = [c["name"] for c in cfg_columns]
        unique_columns = [c["name"] for c in cfg_columns if c.get("unique", False)]

        # Read datasources
        dataframes = []
        for ds in datasources:
            df = ds["dataframe"]

            # Build column map for this datasource
            column_map = {}
            columns_found = []
            for col in ds["column_spec"]:
                if "header" in col:
                    column_map[col["header"]["name"]] = col["name"]
                    columns_found.append(col["name"])
            df.rename(columns=column_map, inplace=True)

            # Add missing columns
            for c in columns:
                if c not in columns_found:
                    logger.debug("Adding missing column {}".format(c))
                    df[c] = np.nan

            # Sort columns
            df = df[columns].copy()

            # Add sort key
            df["sort_key"] = ds["sort_key"]

            # See if we need special type handling on any columns
            for col in ds["column_spec"]:
                col_type = col.get("type")
                col_name = col["name"]
                if col_type == "date":
                    size_before = df[col_name].count()

                    after_col = col_name + "_after_convert"
                    df[after_col] = pd.to_datetime(df[col_name], dayfirst=True, errors="coerce")

                    error_values = df[~df[col_name].isnull() & df[after_col].isnull()] \
                        [col_name].values.tolist()
                    error_values = [{"sourcename": ds["sourcename"],
                                     "sheetname": ds["sheetname"],
                                     "column": col_name,
                                     "column_type": col_type,
                                     "original": v
                                     } for v in error_values]

                    error_report += error_values

                    df[col_name] = df[after_col]
                    df[col_name] = df[col_name].dt.date
                    size_after = df[col_name].count()
                    if size_before != size_after:
                        logger.warn("Removed {} dates from {} in {} due to invalid format."
                                    .format(size_before - size_after, ds["sheetname"], ds["sourcename"]))

            dataframes.append(df)

        if len(dataframes) > 0:
            logger.debug("Merging {} dataframes for {} with the following unique columns: {}" \
                         .format(len(dataframes), ds_key, unique_columns))

            df_lengths = []
            for df in dataframes:
                df_lengths.append(df.shape[0])

            dataframes = pd.concat(dataframes)
            len_before = dataframes.shape[0]

            dataframes.sort_values("sort_key", inplace=True)

            for c in unique_columns:
                logger.debug("Column {} has {} unique values".format(c, len(dataframes[c].unique())))

            dataframes = dataframes.groupby(unique_columns).last().reset_index()
            dataframes = dataframes[columns]

            output_dataframes[ds_key] = dataframes

            len_after = dataframes.shape[0]

            if len_after <= np.min(df_lengths):
                logger.warning("Low number of rows after deduplication - could indicate a problem. " +
                               "Before: {} After: {}".format(len_before, len_after))

            logger.debug("Deduplicated {} rows giving {} unique rows  for {}".format(len_before, len_after, ds_key))

    if error_report_filename is not None and len(error_report) > 0:
        error_report = pd.DataFrame(error_report)
        error_report = error_report[["sourcename", "sheetname", "column", "column_type", "original"]]
        error_report.to_excel(error_report_filename, index=False)
        logger.info("Wrote column error report to {}".format(error_report_filename))

    return output_dataframes


def write_dataframes(dataframes, datasources, output_file, **args):
    import pandas as pd
    cfg_datasources = datasources

    logger.info("Writing to {}".format(output_file))
    writer = pd.ExcelWriter(output_file, engine='xlsxwriter')

    for ds_key, ds in cfg_datasources.items():
        df = dataframes.get(ds_key)
        if df is None:
            df = pd.DataFrame(columns=[x["name"] for x in ds["columns"]])
        logger.info("Writing {} rows to {}".format(df.shape[0], ds_key))
        df.to_excel(writer, sheet_name=ds["name"], index=False)

    writer.save()


def parse_input_config(value, wrap_in_list=True):
    """
    Parses an value for `input_files`. Valid values are a single item, or a list of one of the following:
      * A string - interpreted as the pattern
      * An object - interpreted as the input_config object with root and include values

      Returns a list with the input objects
    """

    if isinstance(value, str):
        value = dict(include=value)

    if isinstance(value, dict):
        if "root" not in value:
            value["root"] = os.getcwd()

    if wrap_in_list:
        if not isinstance(value, list):
            value = [value]

        value = [parse_input_config(v, False) for v in value]

    return value


def merge(config):
    # Make sure we have input_files configured
    if "input_files" not in config:
        raise ConfigError("No 'input_files' configured.")

    config["input_files"] = parse_input_config(config["input_files"])

    # First we scan the input files section for all inputs
    files = []
    for cfg_input in config["input_files"]:
        files += find_input_files(**cfg_input)

    logger.info("Found {} candidate input files".format(len(files)))

    # We then scan the input files for data sources
    data_sources = []
    for file in files:
        data_sources += find_worksheets(**file)

    logger.info("Found {} candidate data sources".format(len(data_sources)))

    # Read Global Datasource Configuration
    datasources_config = init_datasource_config(config["datasources"])

    # Match datasources based on configuration
    datasources_config_matched = match_datasources(data_sources, datasources_config)

    # Limit to only matched datasources
    valid_datasources = [d for d in datasources_config_matched if "source_key" in d]

    # Initialise column configuration for all valid datasources
    valid_datasources = init_all_column_config(valid_datasources)

    # Match headers to column configuration
    valid_datasources = match_columns(valid_datasources)

    # Write column report
    if config.get("column_report_filename") is not None:
        column_report(valid_datasources, datasources_config_matched, config["column_report_filename"])

    # Load dataframes
    valid_datasources = load_dataframes(valid_datasources, **config)

    # Merge sources into single view
    valid_datasources = merge_datasources(valid_datasources, **config)

    # Write final output
    write_dataframes(valid_datasources, **config)

    return valid_datasources
